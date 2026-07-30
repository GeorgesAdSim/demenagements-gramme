#!/usr/bin/env python3
"""
Remplit distance_km et temps_min dans distances-communes-gramme.xlsx
via Nominatim (géocodage) et OSRM (itinéraire routier).

À exécuter depuis une machine ayant accès à Internet — le conteneur de la
session Cowork n'atteint ni Nominatim ni les portails open data.

    pip install openpyxl requests
    python scripts/remplir-distances.py --calibrer      # à lancer EN PREMIER
    python scripts/remplir-distances.py --remplir

Deux étapes volontairement séparées. `--calibrer` ne modifie rien : il
remesure les communes dont tu connais déjà la distance et affiche l'écart
avec ta valeur. Si l'écart est important, les 80 autres lignes seraient
fausses de la même façon — mieux vaut le savoir avant de les écrire.
"""

import argparse
import os
import sys
import time
import urllib.parse

try:
    import requests
    import openpyxl
except ImportError as e:
    sys.exit(f"dépendance manquante : {e.name} — lance `pip install openpyxl requests`")

FEUILLE = 'Distances'

# Nominatim exige un User-Agent identifiant l'application ET un contact réel.
# Sans lui, les requêtes sont refusées ou limitées sans avertissement.
UA = 'DemenagementsGrammeSEO/1.0 (contact@demenagements-gramme.be)'

# Politique d'usage de Nominatim : une requête par seconde maximum. On prend
# une marge, le géocodage de 80 communes prend environ trois minutes.
PAUSE = 1.5
DELAI = 20

# Colonnes, indexées à partir de 1 (openpyxl).
COL = {
    'slug': 2, 'commune': 3, 'depart': 5, 'arrivee': 6,
    'distance': 7, 'temps': 8, 'remarque': 11,
}


def geocoder(adresse):
    """Renvoie (lat, lon, libellé retenu) ou (None, None, None)."""
    url = ('https://nominatim.openstreetmap.org/search?'
           + urllib.parse.urlencode({'q': adresse, 'format': 'json', 'limit': 1,
                                     'countrycodes': 'be'}))
    try:
        r = requests.get(url, headers={'User-Agent': UA}, timeout=DELAI)
        r.raise_for_status()
        res = r.json()
    except Exception as e:
        print(f'    géocodage impossible ({type(e).__name__})')
        return None, None, None
    if not res:
        print('    aucun résultat de géocodage')
        return None, None, None
    return float(res[0]['lat']), float(res[0]['lon']), res[0].get('display_name', '')


def itineraire(lat1, lon1, lat2, lon2):
    """Renvoie (km entier, minutes entières, mètres de recalage du point d'arrivée)."""
    url = (f'https://router.project-osrm.org/route/v1/driving/'
           f'{lon1},{lat1};{lon2},{lat2}?overview=false')
    try:
        r = requests.get(url, timeout=DELAI)
        r.raise_for_status()
        d = r.json()
    except Exception as e:
        print(f'    itinéraire impossible ({type(e).__name__})')
        return None, None, None
    if d.get('code') != 'Ok' or not d.get('routes'):
        print(f"    OSRM a répondu « {d.get('code')} »")
        return None, None, None
    route = d['routes'][0]
    # Entiers : les trois valeurs déjà validées le sont, et le site annonce
    # « environ X km ». Mélanger 8 et 33,1 dans la même colonne n'apporte
    # qu'une fausse précision.
    km = round(route['distance'] / 1000)
    minutes = round(route['duration'] / 60)
    recalage = max(w.get('distance', 0) for w in d.get('waypoints', [{}]))
    return km, minutes, recalage


def charger(FICHIER):
    if not os.path.exists(FICHIER):
        sys.exit(f'fichier introuvable : {FICHIER}\n'
                 f'Lance le script depuis le dossier contenant le .xlsx, ou passe --fichier <chemin>.')
    wb = openpyxl.load_workbook(FICHIER)
    if FEUILLE not in wb.sheetnames:
        sys.exit(f'feuille « {FEUILLE} » absente de {FICHIER}')
    return wb, wb[FEUILLE]


def lignes(ws):
    for r in range(2, ws.max_row + 1):
        slug = ws.cell(row=r, column=COL['slug']).value
        if slug:
            yield r, slug


def calibrer(FICHIER):
    """Remesure les lignes déjà renseignées et affiche l'écart. N'écrit rien."""
    _, ws = charger(FICHIER)
    depart = ws.cell(row=2, column=COL['depart']).value
    print(f'Départ : {depart}')
    dlat, dlon, libelle = geocoder(depart)
    if dlat is None:
        sys.exit('le dépôt n\'a pas pu être géocodé — tout le reste en dépendrait.')
    print(f'  géocodé sur : {libelle}\n  {dlat}, {dlon}\n')
    time.sleep(PAUSE)

    connues = []
    for r, slug in lignes(ws):
        km_ref = ws.cell(row=r, column=COL['distance']).value
        min_ref = ws.cell(row=r, column=COL['temps']).value
        if km_ref in (None, 0):
            continue
        connues.append((r, slug, km_ref, min_ref))

    if not connues:
        sys.exit('aucune ligne déjà renseignée — rien à calibrer.')

    print(f'{"commune":22} {"réf.":>10} {"OSRM":>10} {"écart":>9}   recalage')
    print('-' * 68)
    ecarts = []
    for r, slug, km_ref, min_ref in connues:
        arrivee = ws.cell(row=r, column=COL['arrivee']).value
        lat, lon, _ = geocoder(arrivee)
        time.sleep(PAUSE)
        if lat is None:
            continue
        km, minutes, recalage = itineraire(dlat, dlon, lat, lon)
        time.sleep(PAUSE)
        if km is None:
            continue
        ecart = (km - km_ref) / km_ref * 100
        ecarts.append(abs(ecart))
        print(f'{slug:22} {km_ref:>7} km {km:>7} km {ecart:>+8.0f}%   {recalage:>5.0f} m')

    if not ecarts:
        sys.exit('\naucune mesure aboutie — vérifie ta connexion.')

    pire = max(ecarts)
    print('-' * 68)
    print(f'\nÉcart maximal : {pire:.0f}%')
    if pire <= 10:
        print('→ Méthode cohérente avec tes relevés. Tu peux lancer --remplir.')
    else:
        print('→ Écart trop important pour publier ces chiffres tels quels.')
        print('  Soit le géocodage tombe à côté du centre de la commune,')
        print('  soit l\'une de tes valeurs de référence est à revoir.')
        print('  Vérifie les lignes ci-dessus dans Google Maps avant --remplir.')
        print('  Le recalage indique de combien OSRM a déplacé le point pour')
        print('  rejoindre une route : au-delà de 500 m, le point est mal placé.')


def remplir(FICHIER):
    """Complète les cellules vides. Préserve la mise en forme et le second onglet."""
    wb, ws = charger(FICHIER)
    depart = ws.cell(row=2, column=COL['depart']).value
    dlat, dlon, libelle = geocoder(depart)
    if dlat is None:
        sys.exit('le dépôt n\'a pas pu être géocodé.')
    print(f'Départ géocodé sur : {libelle}\n')
    time.sleep(PAUSE)

    faits, echecs = 0, []
    for r, slug in lignes(ws):
        if ws.cell(row=r, column=COL['distance']).value is not None:
            continue
        arrivee = ws.cell(row=r, column=COL['arrivee']).value
        print(f'  {slug} …')
        lat, lon, libelle_arr = geocoder(arrivee)
        time.sleep(PAUSE)
        if lat is None:
            echecs.append((slug, 'géocodage'))
            continue
        km, minutes, recalage = itineraire(dlat, dlon, lat, lon)
        time.sleep(PAUSE)
        if km is None:
            echecs.append((slug, 'itinéraire'))
            continue

        ws.cell(row=r, column=COL['distance']).value = km
        ws.cell(row=r, column=COL['temps']).value = minutes
        note = f'OSRM/OSM, recalage {recalage:.0f} m'
        if recalage > 500:
            note += ' — À VÉRIFIER, point éloigné de toute route'
        ws.cell(row=r, column=COL['remarque']).value = note
        faits += 1

    # Enregistrement avec openpyxl, et non pandas.to_excel : celui-ci
    # reconstruit le classeur et perdrait l'onglet « Mode opératoire », les
    # remplissages jaunes des cellules à compléter, le style d'en-tête et le
    # filtre automatique.
    wb.save(FICHIER)
    print(f'\n{faits} ligne(s) complétée(s) dans {FICHIER}.')
    if echecs:
        print(f'{len(echecs)} échec(s), à faire à la main :')
        for slug, etape in echecs:
            print(f'  · {slug} ({etape})')


if __name__ == '__main__':
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--calibrer', action='store_true', help='remesure les lignes connues, n\'écrit rien')
    ap.add_argument('--remplir', action='store_true', help='complète les cellules vides')
    ap.add_argument('--fichier', default='distances-communes-gramme.xlsx',
                    help='chemin du classeur (défaut : dans le dossier courant)')
    a = ap.parse_args()
    if a.calibrer:
        calibrer(a.fichier)
    elif a.remplir:
        remplir(a.fichier)
    else:
        ap.print_help()
